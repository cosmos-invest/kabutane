from __future__ import annotations

import argparse
import concurrent.futures
import csv
import hashlib
import io
import json
import math
import re
import time
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
JPX_ISSUES_URL = "https://www.jpx.co.jp/equities/products/etfs/issues/01.html"
JPX_DATABASE_URL = (
    "https://www.jpx.co.jp/equities/products/etfs/investors/"
    "tvdivq0000005cdd-att/nlsgeu000000vx9t.xlsx"
)
OUTPUT_ROOT = ROOT / "data" / "premium" / "etf-pcf"
CODE_RE = re.compile(r"^(?:\d{4}|[0-9]{3}[A-Z])$")
STANDARD_FIELDS = {
    "etf code",
    "etf name",
    "fund cash component",
    "shares outstanding",
    "fund date",
}
SOLACTIVE_EXCEPTIONS = {"2858", "179A", "180A", "380A", "404A", "502A", "563A", "576A"}
ICE_ZIP_LIST_URL = "https://inav.ice.com/pcf-download/listOfZips"
ICE_ALL_ZIP_URL = "https://inav.ice.com/pcf-download/all/{name}"
HISTORY_RETENTION = 45
ICE_ZIP_RE = re.compile(r"all_pcf_(\d{8})\.zip", re.IGNORECASE)
ICE_MEMBER_RE = re.compile(r"(?:^|/)([0-9]{4}|[0-9]{3}[A-Z])tsepcf_", re.IGNORECASE)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def finite(value: Any) -> float | None:
    if value in (None, "", "-", "nan", "NaN"):
        return None
    try:
        number = float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def compact_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u3000", " ").split())


def normalize_code(value: Any) -> str:
    text = compact_text(value).upper()
    return text if CODE_RE.fullmatch(text) else ""


def clean_active_name(value: Any) -> str:
    text = compact_text(value)
    text = re.sub(r"\s*iNAV\s*", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s*アクティブ運用型(?:\s*\((?:特化型|デリバティブ)\))?\s*", " ", text)
    return compact_text(text)


def write_json_if_changed(path: Path, payload: Any) -> bool:
    text = json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n"
    if path.exists() and path.read_text(encoding="utf-8") == text:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    return True


def load_json(path: Path, default: Any) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def download(url: str, timeout: int = 30, retries: int = 2) -> bytes:
    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            request = Request(
                url,
                headers={
                    "User-Agent": "kabutane-etf-pcf/1.0 (+https://github.com/cosmos-invest/kabutane)",
                    "Accept": "text/csv,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/html,*/*",
                },
            )
            with urlopen(request, timeout=timeout) as response:
                return response.read()
        except (HTTPError, URLError, TimeoutError, OSError) as exc:
            last_error = exc
            if attempt < retries:
                time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"download failed: {url}: {last_error}")


@dataclass
class TableCell:
    text: list[str] = field(default_factory=list)
    links: list[str] = field(default_factory=list)


class JpxTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[TableCell]] = []
        self.row: list[TableCell] | None = None
        self.cell: TableCell | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self.row = []
        elif tag in {"td", "th"} and self.row is not None:
            self.cell = TableCell()
            self.row.append(self.cell)
        elif tag == "a" and self.cell is not None:
            href = dict(attrs).get("href")
            if href:
                self.cell.links.append(href)

    def handle_data(self, data: str) -> None:
        if self.cell is not None and data.strip():
            self.cell.text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"}:
            self.cell = None
        elif tag == "tr" and self.row is not None:
            if self.row:
                self.rows.append(self.row)
            self.row = None
            self.cell = None


def discover_active_etfs(html_bytes: bytes) -> list[dict[str, Any]]:
    text = html_bytes.decode("utf-8", errors="replace")
    parser = JpxTableParser()
    parser.feed(text)
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for cells in parser.rows:
        values = [compact_text("".join(cell.text)) for cell in cells]
        joined = " ".join(values)
        if "アクティブ運用型" not in joined:
            continue
        code_index = next((i for i, value in enumerate(values) if normalize_code(value)), None)
        if code_index is None:
            continue
        code = normalize_code(values[code_index])
        if not code or code in seen:
            continue
        name = values[code_index + 1] if code_index + 1 < len(values) else code
        sponsor = values[code_index + 2] if code_index + 2 < len(values) else ""
        links = [href for cell in cells for href in cell.links]
        provider_hint = ""
        if any("inav.ice.com" in href or "factsetdigitalsolutions.com" in href for href in links):
            provider_hint = "ICE"
        elif any("ihsmarkit.com" in href or "spglobal.com" in href for href in links):
            provider_hint = "S&P Global"
        elif "Global X" in sponsor and code not in SOLACTIVE_EXCEPTIONS:
            provider_hint = "Solactive"
        records.append(
            {
                "code": code,
                "name": clean_active_name(name),
                "sponsor": sponsor,
                "specialized": "特化型" in joined,
                "derivative": "デリバティブ" in joined,
                "provider_hint": provider_hint,
                "links": links,
            }
        )
        seen.add(code)
    return records


def pcf_urls_from_workbook(workbook_bytes: bytes) -> tuple[dict[str, dict[str, Any]], str | None]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    header_row = None
    headers: list[str] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        values = [compact_text(value) for value in row]
        if "コード" in values and any("PCF掲載URL" in value for value in values):
            header_row = row_index
            headers = values
            break
    if header_row is None:
        raise ValueError("JPX ETF database header was not found")
    code_idx = headers.index("コード")
    name_idx = headers.index("名称")
    category_idx = headers.index("カテゴリー")
    sponsor_idx = headers.index("管理会社")
    url_idx = next(i for i, value in enumerate(headers) if value.startswith("PCF掲載URL") and "日本語" in value)
    frequency_idx = next(i for i, value in enumerate(headers) if value.startswith("PCF 更新頻度") or value.startswith("PCF更新頻度"))
    records: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        code = normalize_code(row[code_idx] if code_idx < len(row) else "")
        if not code:
            continue
        records[code] = {
            "name": compact_text(row[name_idx]),
            "category": compact_text(row[category_idx]),
            "sponsor": compact_text(row[sponsor_idx]),
            "url": compact_text(row[url_idx]),
            "frequency": compact_text(row[frequency_idx]),
        }
    title = compact_text(sheet.cell(1, 2).value)
    match = re.search(r"(\d{4})年(\d{1,2})月", title)
    as_of = f"{match.group(1)}-{int(match.group(2)):02d}" if match else None
    return records, as_of


def resolve_pcf_url(record: dict[str, Any], database_record: dict[str, Any] | None) -> tuple[str, str]:
    database_url = compact_text((database_record or {}).get("url"))
    if database_url.startswith("http"):
        if "inav.ice.com" in database_url:
            return database_url, "ICE"
        if "ihsmarkit.com" in database_url or "spglobal.com" in database_url:
            return database_url, "S&P Global"
        if "solactive.com" in database_url:
            return database_url, "Solactive"
        return database_url, "運用会社"
    code = record["code"]
    hint = record.get("provider_hint")
    if hint == "ICE":
        return f"https://inav.ice.com/pcf-download/{code}.csv", "ICE"
    if hint == "S&P Global":
        return f"https://api.ebs.ihsmarkit.com/inav/getfile?filename={code}.csv", "S&P Global"
    if hint == "Solactive":
        return f"https://legacy2.solactive.com/downloads/etfservices/tse-pcf/single/{code}.csv", "Solactive"
    return "", "未確認"


def decode_csv(content: bytes) -> str:
    if content.lstrip().lower().startswith((b"<html", b"<!doctype")):
        raise ValueError("PCF download service returned an HTML notice")
    for encoding in ("utf-8-sig", "cp932", "shift_jis"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("unsupported PCF text encoding")


def parse_pcf(content: bytes, source_url: str = "", provider: str = "") -> dict[str, Any]:
    rows = list(csv.reader(io.StringIO(decode_csv(content))))
    metadata_header_index = next(
        (
            i
            for i, row in enumerate(rows)
            if STANDARD_FIELDS.issubset({compact_text(value).lower() for value in row})
        ),
        None,
    )
    if metadata_header_index is None or metadata_header_index + 1 >= len(rows):
        raise ValueError("PCF metadata header was not found")
    metadata_header = [compact_text(value).lower() for value in rows[metadata_header_index]]
    metadata_values = rows[metadata_header_index + 1]
    metadata = {
        key: compact_text(metadata_values[index]) if index < len(metadata_values) else ""
        for index, key in enumerate(metadata_header)
        if key
    }
    holdings_header_index = next(
        (
            i
            for i, row in enumerate(rows[metadata_header_index + 2 :], start=metadata_header_index + 2)
            if {"code", "name", "shares amount"}.issubset({compact_text(value).lower() for value in row})
        ),
        None,
    )
    if holdings_header_index is None:
        raise ValueError("PCF holdings header was not found")
    holding_headers = [compact_text(value).lower() for value in rows[holdings_header_index]]
    fund_date_raw = re.sub(r"\D", "", metadata.get("fund date", ""))
    fund_date = (
        f"{fund_date_raw[:4]}-{fund_date_raw[4:6]}-{fund_date_raw[6:8]}"
        if len(fund_date_raw) == 8
        else compact_text(metadata.get("fund date"))
    )
    shares_outstanding = finite(metadata.get("shares outstanding"))
    if not shares_outstanding or shares_outstanding <= 0:
        raise ValueError("PCF shares outstanding is missing or invalid")
    holdings: list[dict[str, Any]] = []
    for raw in rows[holdings_header_index + 1 :]:
        if not any(compact_text(value) for value in raw):
            continue
        item = {
            key: compact_text(raw[index]) if index < len(raw) else ""
            for index, key in enumerate(holding_headers)
            if key
        }
        amount = finite(item.get("shares amount"))
        price = finite(item.get("stock price"))
        if amount is None:
            continue
        code = normalize_code(item.get("code"))
        isin = compact_text(item.get("isin"))
        name = compact_text(item.get("name"))
        stable_id = code or isin or f"{name}|{compact_text(item.get('currency'))}|{compact_text(item.get('exchange'))}"
        holdings.append(
            {
                "id": stable_id,
                "code": code or None,
                "name": name,
                "isin": isin or None,
                "exchange": compact_text(item.get("exchange")) or None,
                "currency": compact_text(item.get("currency")) or None,
                "shares_amount": amount,
                "stock_price": price,
                "market_value": round(amount * price, 6) if price is not None else None,
                "units_per_million": round(amount / shares_outstanding * 1_000_000, 8),
            }
        )
    if not holdings:
        raise ValueError("PCF contains no holdings")
    payload = {
        "schema_version": 1,
        "kind": "kabutane_etf_pcf_snapshot",
        "etf_code": normalize_code(metadata.get("etf code")),
        "etf_name": compact_text(metadata.get("etf name")),
        "fund_date": fund_date,
        "fund_cash_component": finite(metadata.get("fund cash component")),
        "shares_outstanding": shares_outstanding,
        "source_url": source_url,
        "source_provider": provider,
        "source_sha256": hashlib.sha256(content).hexdigest(),
        "retrieved_at": utc_now(),
        "holdings": holdings,
    }
    return payload


def load_stock_names(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return {normalize_code(row.get("code")): compact_text(row.get("name")) for row in csv.DictReader(handle) if normalize_code(row.get("code"))}


def enrich_snapshot(snapshot: dict[str, Any], stock_names: dict[str, str]) -> dict[str, Any]:
    holdings = snapshot.get("holdings") or []
    total_jpy_value = sum(
        float(row.get("market_value") or 0)
        for row in holdings
        if row.get("currency") == "JPY" and row.get("market_value") is not None
    )
    cash = finite(snapshot.get("fund_cash_component")) or 0.0
    denominator = total_jpy_value + cash if total_jpy_value + cash > 0 else total_jpy_value
    domestic_value = 0.0
    reit_value = 0.0
    for row in holdings:
        code = normalize_code(row.get("code"))
        if code and code in stock_names:
            row["name_ja"] = stock_names[code]
        market_value = finite(row.get("market_value")) or 0.0
        if row.get("currency") == "JPY" and code:
            domestic_value += market_value
            if code.startswith(("134", "1476", "1597", "1660")) or "REIT" in str(row.get("name") or "").upper():
                reit_value += market_value
        row["weight_pct"] = round(market_value / denominator * 100, 5) if denominator > 0 and row.get("currency") == "JPY" else None
    domestic_ratio = domestic_value / denominator if denominator > 0 else 0.0
    snapshot["domestic_equity"] = domestic_ratio >= 0.5 and (reit_value / domestic_value if domestic_value else 0) < 0.5
    snapshot["asset_scope"] = "国内株" if snapshot["domestic_equity"] else "その他資産"
    snapshot["cash_weight_pct"] = round(cash / denominator * 100, 4) if denominator > 0 else None
    snapshot["holding_count"] = len(holdings)
    return snapshot


def classify_change(current: dict[str, Any] | None, previous: dict[str, Any] | None) -> dict[str, Any] | None:
    if current is None and previous is None:
        return None
    base = current or previous or {}
    if previous is None:
        kind = "NEW"
    elif current is None:
        kind = "REMOVED"
    else:
        current_units = finite(current.get("units_per_million")) or 0.0
        previous_units = finite(previous.get("units_per_million")) or 0.0
        delta = current_units - previous_units
        delta_pct = delta / abs(previous_units) * 100 if previous_units else None
        weight_delta = (
            float(current["weight_pct"]) - float(previous["weight_pct"])
            if current.get("weight_pct") is not None and previous.get("weight_pct") is not None
            else None
        )
        # Weight alone moves when the stock price changes.  A manager-action
        # candidate therefore requires a change in quantity per ETF unit.
        material = delta_pct is not None and abs(delta_pct) >= 1.0
        if not material:
            return None
        kind = "INCREASE" if delta > 0 else "DECREASE"
    current_units = finite((current or {}).get("units_per_million"))
    previous_units = finite((previous or {}).get("units_per_million"))
    units_delta = (current_units - previous_units) if current_units is not None and previous_units is not None else None
    units_delta_pct = units_delta / abs(previous_units) * 100 if units_delta is not None and previous_units else None
    weight_current = finite((current or {}).get("weight_pct"))
    weight_previous = finite((previous or {}).get("weight_pct"))
    weight_delta = weight_current - weight_previous if weight_current is not None and weight_previous is not None else None
    value_current = finite((current or {}).get("market_value"))
    value_previous = finite((previous or {}).get("market_value"))
    value_per_unit_current = value_current / current_units if value_current is not None and current_units else None
    value_per_unit_previous = value_previous / previous_units if value_previous is not None and previous_units else None
    price_ratio = value_per_unit_current / value_per_unit_previous if value_per_unit_current and value_per_unit_previous else None
    quantity_ratio = current_units / previous_units if current_units and previous_units else None
    possible_corporate_action = bool(
        quantity_ratio
        and price_ratio
        and (quantity_ratio >= 1.5 or quantity_ratio <= 0.67)
        and abs(quantity_ratio * price_ratio - 1) <= 0.15
    )
    return {
        "id": base.get("id"),
        "code": base.get("code"),
        "name": base.get("name"),
        "name_ja": base.get("name_ja"),
        "kind": kind,
        "units_per_million": current_units,
        "units_delta": round(units_delta, 8) if units_delta is not None else None,
        "units_delta_pct": round(units_delta_pct, 3) if units_delta_pct is not None else None,
        "weight_pct": weight_current,
        "weight_delta_pct_point": round(weight_delta, 5) if weight_delta is not None else None,
        "possible_corporate_action": possible_corporate_action,
    }


def compare_snapshots(current: dict[str, Any], previous: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not previous or previous.get("fund_date") == current.get("fund_date"):
        return []
    current_map = {str(row.get("id")): row for row in current.get("holdings") or []}
    previous_map = {str(row.get("id")): row for row in previous.get("holdings") or []}
    changes = [
        change
        for key in sorted(current_map.keys() | previous_map.keys())
        if (change := classify_change(current_map.get(key), previous_map.get(key))) is not None
    ]
    order = {"NEW": 0, "INCREASE": 1, "DECREASE": 2, "REMOVED": 3}
    changes.sort(
        key=lambda row: (
            order.get(str(row.get("kind")), 9),
            -abs(float(row.get("weight_delta_pct_point") or row.get("units_delta_pct") or 0)),
            str(row.get("code") or row.get("name") or ""),
        )
    )
    return changes


def ice_archive_names(content: bytes, limit: int = 8) -> list[str]:
    """Return the latest official ICE daily archive names from either JSON or text."""
    names = {f"all_pcf_{match}.zip" for match in ICE_ZIP_RE.findall(content.decode("utf-8", errors="replace"))}
    return sorted(names, reverse=True)[:limit]


def snapshots_from_ice_archive(content: bytes, active_codes: set[str] | None = None) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        for member in archive.namelist():
            match = ICE_MEMBER_RE.search(member)
            if not match:
                continue
            code = normalize_code(match.group(1))
            if active_codes is not None and code not in active_codes:
                continue
            try:
                snapshot = parse_pcf(archive.read(member), ICE_ALL_ZIP_URL.format(name=Path(member).name), "ICE")
            except (KeyError, ValueError):
                continue
            snapshot["etf_code"] = snapshot.get("etf_code") or code
            result[code] = snapshot
    return result


def compact_snapshot(snapshot: dict[str, Any]) -> dict[str, Any]:
    keys = (
        "schema_version", "kind", "etf_code", "etf_name", "fund_date", "fund_cash_component",
        "shares_outstanding", "source_url", "source_provider", "source_sha256", "retrieved_at",
        "holdings", "domestic_equity", "asset_scope", "cash_weight_pct", "holding_count",
        "sponsor", "specialized", "derivative",
    )
    return {key: snapshot.get(key) for key in keys if key in snapshot}


def merge_snapshot_history(existing: Iterable[dict[str, Any]], additions: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    by_date = {
        str(snapshot.get("fund_date")): compact_snapshot(snapshot)
        for snapshot in [*existing, *additions]
        if snapshot.get("fund_date")
    }
    return [by_date[date] for date in sorted(by_date)[-HISTORY_RETENTION:]]


def period_baseline(history: list[dict[str, Any]], current_date: str, days: int) -> dict[str, Any] | None:
    try:
        target = datetime.strptime(current_date, "%Y-%m-%d").date() - timedelta(days=days)
    except (TypeError, ValueError):
        return None
    eligible = []
    for snapshot in history:
        try:
            snapshot_date = datetime.strptime(str(snapshot.get("fund_date")), "%Y-%m-%d").date()
        except (TypeError, ValueError):
            continue
        if snapshot_date <= target:
            eligible.append((snapshot_date, snapshot))
    return max(eligible, key=lambda item: item[0])[1] if eligible else None


def period_comparison(current: dict[str, Any], history: list[dict[str, Any]], days: int) -> dict[str, Any]:
    baseline = period_baseline(history, str(current.get("fund_date") or ""), days)
    if baseline is None:
        return {"status": "collecting", "baseline_date": None, "changes": [], "change_counts": {}, "material_change_count": 0}
    changes = compare_snapshots(current, baseline)
    counts = {kind: sum(row.get("kind") == kind for row in changes) for kind in ("NEW", "INCREASE", "DECREASE", "REMOVED")}
    return {
        "status": "ready",
        "baseline_date": baseline.get("fund_date"),
        "changes": changes[:30],
        "change_counts": counts,
        "material_change_count": len(changes),
    }


def build_periods(current: dict[str, Any], history: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        "day": period_comparison(current, history, 1),
        "week": period_comparison(current, history, 7),
        "month": period_comparison(current, history, 30),
    }


def build_streaks(history: list[dict[str, Any]]) -> list[dict[str, Any]]:
    streaks: dict[str, dict[str, Any]] = {}
    for previous, current in zip(history, history[1:]):
        changes = compare_snapshots(current, previous)
        seen: set[str] = set()
        for change in changes:
            security = str(change.get("code") or change.get("id") or "")
            direction = "UP" if change.get("kind") in {"NEW", "INCREASE"} else "DOWN"
            if not security:
                continue
            seen.add(security)
            old = streaks.get(security)
            count = int(old.get("count") or 0) + 1 if old and old.get("direction") == direction else 1
            streaks[security] = {"code": change.get("code"), "name": change.get("name_ja") or change.get("name"), "direction": direction, "count": count}
        for security in set(streaks) - seen:
            streaks.pop(security, None)
    return sorted((row for row in streaks.values() if row["count"] >= 2), key=lambda row: (-row["count"], str(row.get("code") or "")))[:20]


def build_stock_lookup(funds: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for fund in funds:
        if not fund.get("domestic_equity"):
            continue
        for period, comparison in (fund.get("periods") or {}).items():
            if comparison.get("status") != "ready":
                continue
            for change in comparison.get("changes") or []:
                code = normalize_code(change.get("code"))
                if not code:
                    continue
                row = grouped.setdefault(code, {"code": code, "name": change.get("name_ja") or change.get("name"), "funds": []})
                row["funds"].append({"fund_code": fund.get("code"), "fund_name": fund.get("display_name") or fund.get("name"), "period": period, "kind": change.get("kind"), "units_delta_pct": change.get("units_delta_pct")})
    result = list(grouped.values())
    for row in result:
        row["fund_count"] = len({item["fund_code"] for item in row["funds"]})
    result.sort(key=lambda row: (-int(row["fund_count"]), str(row["code"])))
    return result


def build_sponsor_trends(funds: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for fund in funds:
        sponsor = compact_text(fund.get("sponsor")) or "運用会社未確認"
        row = grouped.setdefault(sponsor, {"sponsor": sponsor, "fund_count": 0, "periods": {}})
        row["fund_count"] += 1
        for period, comparison in (fund.get("periods") or {}).items():
            bucket = row["periods"].setdefault(period, {"up": 0, "down": 0, "status": "collecting"})
            if comparison.get("status") == "ready":
                bucket["status"] = "ready"
                for change in comparison.get("changes") or []:
                    bucket["up" if change.get("kind") in {"NEW", "INCREASE"} else "down"] += 1
    return sorted(grouped.values(), key=lambda row: (-int(row["fund_count"]), row["sponsor"]))


def summarize_snapshot(snapshot: dict[str, Any], changes: list[dict[str, Any]], baseline: bool) -> dict[str, Any]:
    counts = {kind: sum(row.get("kind") == kind for row in changes) for kind in ("NEW", "INCREASE", "DECREASE", "REMOVED")}
    top_holdings = sorted(
        (row for row in snapshot.get("holdings") or [] if row.get("weight_pct") is not None),
        key=lambda row: -float(row.get("weight_pct") or 0),
    )[:10]
    return {
        "code": snapshot.get("etf_code"),
        "name": snapshot.get("etf_name"),
        "fund_date": snapshot.get("fund_date"),
        "provider": snapshot.get("source_provider"),
        "source_url": snapshot.get("source_url"),
        "asset_scope": snapshot.get("asset_scope"),
        "domestic_equity": snapshot.get("domestic_equity") is True,
        "holding_count": snapshot.get("holding_count"),
        "cash_weight_pct": snapshot.get("cash_weight_pct"),
        "baseline": baseline,
        "change_counts": counts,
        "material_change_count": len(changes),
        "changes": changes[:30],
        "top_holdings": [
            {key: row.get(key) for key in ("code", "name", "name_ja", "weight_pct")}
            for row in top_holdings
        ],
    }


def build_common_changes(funds: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for fund in funds:
        if not fund.get("domestic_equity"):
            continue
        for change in fund.get("changes") or []:
            direction = "UP" if change.get("kind") in {"NEW", "INCREASE"} else "DOWN"
            key = (str(change.get("code") or change.get("name") or ""), direction)
            grouped[key].append({"fund_code": fund.get("code"), "fund_name": fund.get("name"), "kind": change.get("kind")})
    result = []
    for (security, direction), rows in grouped.items():
        if security and len(rows) >= 2:
            first_change = next(
                change
                for fund in funds
                for change in fund.get("changes") or []
                if str(change.get("code") or change.get("name") or "") == security
            )
            result.append(
                {
                    "security": security,
                    "code": first_change.get("code"),
                    "name": first_change.get("name_ja") or first_change.get("name"),
                    "direction": direction,
                    "fund_count": len(rows),
                    "funds": rows,
                }
            )
    result.sort(key=lambda row: (-int(row["fund_count"]), str(row.get("code") or row.get("name") or "")))
    return result[:30]


def append_history(path: Path, summary: dict[str, Any]) -> None:
    compact = {
        "fund_date": summary.get("fund_date"),
        "material_change_count": summary.get("material_change_count"),
        "change_counts": summary.get("change_counts"),
        "cash_weight_pct": summary.get("cash_weight_pct"),
        "holding_count": summary.get("holding_count"),
    }
    existing: list[dict[str, Any]] = []
    if path.exists():
        for line in path.read_text(encoding="utf-8").splitlines():
            try:
                existing.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    by_date = {str(row.get("fund_date")): row for row in existing if row.get("fund_date")}
    by_date[str(compact["fund_date"])] = compact
    lines = [json.dumps(by_date[key], ensure_ascii=False, separators=(",", ":")) for key in sorted(by_date)[-400:]]
    text = "\n".join(lines) + ("\n" if lines else "")
    if not path.exists() or path.read_text(encoding="utf-8") != text:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text, encoding="utf-8")


def validate_payload(payload: dict[str, Any]) -> None:
    if payload.get("kind") != "kabutane_active_etf_pcf":
        raise ValueError("unexpected ETF PCF payload kind")
    summary = payload.get("summary")
    funds = payload.get("funds")
    errors = payload.get("errors")
    if not isinstance(summary, dict) or not isinstance(funds, list) or not isinstance(errors, list):
        raise ValueError("ETF PCF payload requires summary, funds and errors")
    target = summary.get("active_target_count")
    if not isinstance(target, int) or target <= 0:
        raise ValueError("ETF PCF active target count must be positive")
    codes = [str(row.get("code") or "") for row in funds if isinstance(row, dict)]
    if any(not normalize_code(code) for code in codes) or len(codes) != len(set(codes)):
        raise ValueError("ETF PCF fund codes must be unique JPX codes")
    if summary.get("available_count") != len(funds):
        raise ValueError("ETF PCF available count mismatch")
    if summary.get("error_count") != len(errors):
        raise ValueError("ETF PCF error count mismatch")
    for fund in funds:
        if not fund.get("fund_date") or not isinstance(fund.get("changes"), list):
            raise ValueError(f"ETF PCF fund summary is incomplete: {fund.get('code')}")
        periods = fund.get("periods")
        if not isinstance(periods, dict) or set(periods) != {"day", "week", "month"}:
            raise ValueError(f"ETF PCF fund periods are incomplete: {fund.get('code')}")
        if any(value.get("status") not in {"ready", "collecting"} for value in periods.values()):
            raise ValueError(f"ETF PCF period status is invalid: {fund.get('code')}")


def run_update(root: Path = ROOT, issues_bytes: bytes | None = None, database_bytes: bytes | None = None) -> dict[str, Any]:
    output_root = root / "data" / "premium" / "etf-pcf"
    issues_content = issues_bytes if issues_bytes is not None else download(JPX_ISSUES_URL)
    database_content = database_bytes if database_bytes is not None else download(JPX_DATABASE_URL)
    active = discover_active_etfs(issues_content)
    database, database_as_of = pcf_urls_from_workbook(database_content)
    stock_names = load_stock_names(root / "stocks.csv")
    funds: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    unchanged = 0
    updated = 0
    resolved: list[tuple[dict[str, Any], str, str]] = []
    for record in active:
        code = record["code"]
        database_record = database.get(code)
        url, provider = resolve_pcf_url(record, database_record)
        if not url:
            errors.append({"code": code, "name": record.get("name"), "reason": "PCF掲載URLを確認できませんでした"})
            continue
        resolved.append((record, url, provider))

    archive_names: list[str] = []
    archive_error: str | None = None
    archived_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    try:
        archive_names = ice_archive_names(download(ICE_ZIP_LIST_URL, 20, 1), limit=8)
        active_codes = {record["code"] for record in active}
        with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
            future_map = {
                executor.submit(download, ICE_ALL_ZIP_URL.format(name=name), 30, 1): name
                for name in archive_names
            }
            for future in concurrent.futures.as_completed(future_map):
                archive_name = future_map[future]
                try:
                    snapshots = snapshots_from_ice_archive(future.result(), active_codes)
                    for code, snapshot in snapshots.items():
                        snapshot["source_url"] = ICE_ALL_ZIP_URL.format(name=archive_name)
                        archived_by_code[code].append(enrich_snapshot(snapshot, stock_names))
                except Exception as exc:
                    archive_error = f"{archive_name}: {exc}"[:240]
    except Exception as exc:
        archive_error = str(exc)[:240]

    download_results: dict[str, bytes | Exception] = {}
    # Keep the provider load small while avoiding one closed/time-limited
    # endpoint blocking the complete daily update.
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        future_map = {
            executor.submit(download, url, 18, 0): record["code"]
            for record, url, _provider in resolved
        }
        for future in concurrent.futures.as_completed(future_map):
            code = future_map[future]
            try:
                download_results[code] = future.result()
            except Exception as exc:
                download_results[code] = exc

    for record, url, provider in resolved:
        code = record["code"]
        fund_path = output_root / "funds" / f"{code}.json"
        previous = load_json(fund_path, None)
        try:
            content = download_results.get(code)
            if isinstance(content, Exception):
                raise content
            if not isinstance(content, bytes):
                raise RuntimeError("PCF response was not received")
            snapshot = enrich_snapshot(parse_pcf(content, url, provider), stock_names)
            snapshot["etf_code"] = snapshot.get("etf_code") or code
            snapshot["etf_name"] = record.get("name") or snapshot.get("etf_name")
            snapshot["sponsor"] = record.get("sponsor")
            snapshot["specialized"] = record.get("specialized") is True
            snapshot["derivative"] = record.get("derivative") is True
            if previous and previous.get("source_sha256") == snapshot.get("source_sha256"):
                unchanged += 1
                snapshot = previous
                changes = list(previous.get("changes") or [])
                baseline = previous.get("baseline") is True
            else:
                changes = compare_snapshots(snapshot, previous)
                baseline = previous is None
                snapshot["changes"] = changes
                snapshot["baseline"] = baseline
                write_json_if_changed(fund_path, snapshot)
                updated += 1
            summary = summarize_snapshot(snapshot, changes, baseline)
            summary["display_name"] = record.get("name")
            summary["sponsor"] = record.get("sponsor")
            summary["specialized"] = record.get("specialized") is True
            summary["derivative"] = record.get("derivative") is True
            history_path = output_root / "snapshot-history" / f"{code}.json"
            stored_history = load_json(history_path, [])
            archive_snapshots = []
            for archived in archived_by_code.get(code, []):
                archived["etf_code"] = archived.get("etf_code") or code
                archived["etf_name"] = record.get("name") or archived.get("etf_name")
                archived["sponsor"] = record.get("sponsor")
                archived["specialized"] = record.get("specialized") is True
                archived["derivative"] = record.get("derivative") is True
                archive_snapshots.append(archived)
            snapshot_history = merge_snapshot_history(stored_history, [*archive_snapshots, snapshot])
            write_json_if_changed(history_path, snapshot_history)
            summary["periods"] = build_periods(snapshot, snapshot_history)
            summary["streaks"] = build_streaks(snapshot_history)
            summary["history_count"] = len(snapshot_history)
            summary["history_start_date"] = snapshot_history[0].get("fund_date") if snapshot_history else None
            funds.append(summary)
            append_history(output_root / "history" / f"{code}.jsonl", summary)
        except Exception as exc:
            errors.append({"code": code, "name": record.get("name"), "provider": provider, "reason": str(exc)[:240]})
            if previous:
                changes = list(previous.get("changes") or [])
                summary = summarize_snapshot(previous, changes, previous.get("baseline") is True)
                summary["display_name"] = record.get("name")
                summary["sponsor"] = record.get("sponsor")
                summary["stale"] = True
                snapshot_history = merge_snapshot_history(
                    load_json(output_root / "snapshot-history" / f"{code}.json", []),
                    archived_by_code.get(code, []),
                )
                summary["periods"] = build_periods(previous, snapshot_history)
                summary["streaks"] = build_streaks(snapshot_history)
                summary["history_count"] = len(snapshot_history)
                summary["history_start_date"] = snapshot_history[0].get("fund_date") if snapshot_history else None
                funds.append(summary)
    funds.sort(
        key=lambda row: (
            0 if row.get("domestic_equity") else 1,
            -int(row.get("material_change_count") or 0),
            str(row.get("code") or ""),
        )
    )
    latest_dates = sorted({str(row.get("fund_date")) for row in funds if row.get("fund_date")})
    payload = {
        "schema_version": 1,
        "kind": "kabutane_active_etf_pcf",
        "generated_at": utc_now(),
        "source": {
            "jpx_issues_url": JPX_ISSUES_URL,
            "jpx_database_url": JPX_DATABASE_URL,
            "jpx_database_as_of": database_as_of,
            "ice_zip_list_url": ICE_ZIP_LIST_URL,
            "ice_archives": archive_names,
            "ice_archive_error": archive_error,
            "notice": "PCFは前日基準のポートフォリオ情報です。組入変化は売買推奨や当日の市場売買を示しません。設定・解約、企業行動、デリバティブ等の影響を含む場合があります。",
        },
        "summary": {
            "active_target_count": len(active),
            "available_count": len(funds),
            "domestic_equity_count": sum(row.get("domestic_equity") is True for row in funds),
            "updated_count": updated,
            "unchanged_count": unchanged,
            "error_count": len(errors),
            "latest_fund_date": latest_dates[-1] if latest_dates else None,
            "history_ready_count": sum(int(row.get("history_count") or 0) > 1 for row in funds),
            "week_ready_count": sum((row.get("periods") or {}).get("week", {}).get("status") == "ready" for row in funds),
            "month_ready_count": sum((row.get("periods") or {}).get("month", {}).get("status") == "ready" for row in funds),
        },
        "common_changes": build_common_changes(funds),
        "stock_lookup": build_stock_lookup(funds),
        "sponsor_trends": build_sponsor_trends(funds),
        "funds": funds,
        "errors": errors,
    }
    validate_payload(payload)
    write_json_if_changed(output_root / "latest.json", payload)
    write_json_if_changed(
        output_root / "status.json",
        {
            "generated_at": payload["generated_at"],
            **payload["summary"],
            "errors": errors,
        },
    )
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Download and compare TSE active ETF PCF files")
    parser.add_argument("--root", type=Path, default=ROOT)
    parser.add_argument("--strict", action="store_true", help="fail instead of retaining the last successful output")
    args = parser.parse_args()
    root = args.root.resolve()
    try:
        payload = run_update(root)
    except Exception as exc:
        latest = root / "data" / "premium" / "etf-pcf" / "latest.json"
        if args.strict or not latest.exists():
            raise
        print(f"ETF PCF warning: update failed; retaining last successful data: {exc}")
        return
    summary = payload["summary"]
    print(
        "ETF PCF: "
        f"active={summary['active_target_count']} available={summary['available_count']} "
        f"domestic={summary['domestic_equity_count']} updated={summary['updated_count']} "
        f"errors={summary['error_count']} date={summary['latest_fund_date']}"
    )


if __name__ == "__main__":
    main()
